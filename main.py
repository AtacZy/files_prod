"""
ETL-система для сбора, обработки и доставки отчётов
====================================================

Этапы работы:
1. Получение данных из API
2. Парсинг и предобработка
3. Загрузка в PostgreSQL
4. Агрегация статистики
5. Создание Excel-отчёта
6. Отправка по Email
7. Логирование всех этапов

Автор: @AtacZy
Дата: 2026-04-17
Версия: 1.0.0
"""

import os
import glob
import ast
import ssl
import logging
import smtplib
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

import requests
import psycopg2
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Импорт конфигурации
from config import (
    API_URL, API_CLIENT, API_CLIENT_KEY, TIMEZONE_OFFSET,
    DB_NAME, DB_USER, DB_PASSWORD, DB_HOST, DB_PORT,
    EMAIL_SENDER, EMAIL_PASSWORD, EMAIL_RECEIVER, SMTP_SERVER, SMTP_PORT
)


# ============================================================
# КЛАСС 1: ЛОГГЕР
# ============================================================

class Logger:
    """
    Управление логированием.
    
    Возможности:
    - Создание папки logs
    - Запись логов в файл с датой в имени
    - Автоматическое удаление логов старше 3 дней
    - Одновременный вывод в файл и консоль
    """
    
    LOG_DIR = "logs"
    KEEP_DAYS = 3
    
    def __init__(self):
        """Инициализация логгера при создании объекта"""
        self.log_date = datetime.now().strftime("%Y-%m-%d")
        self.log_file = None
        self.logger = None
        self._setup()
    
    def _setup(self):
        """Внутренний метод: настройка логирования"""
        # Создаём папку для логов, если её нет
        if not os.path.exists(self.LOG_DIR):
            os.makedirs(self.LOG_DIR)
        
        self.log_file = f"{self.LOG_DIR}/etl_{self.log_date}.log"
        
        # Настройка формата логов
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s | %(levelname)-8s | %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S',
            handlers=[
                logging.FileHandler(self.log_file, encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        
        self.logger = logging.getLogger(__name__)
        self._cleanup_old_logs()
        
        self.logger.info("=" * 50)
        self.logger.info("ЗАПУСК ETL-СИСТЕМЫ")
        self.logger.info("=" * 50)
        self.logger.info(f"Папка логов: {self.LOG_DIR}/")
        self.logger.info(f"Файл лога: {self.log_file}")
    
    def _cleanup_old_logs(self):
        """Удаление логов старше KEEP_DAYS дней"""
        cutoff_date = datetime.now() - timedelta(days=self.KEEP_DAYS)
        log_files = glob.glob(os.path.join(self.LOG_DIR, "etl_*.log"))
        
        deleted = 0
        for log_file in log_files:
            try:
                filename = os.path.basename(log_file)
                date_str = filename.replace("etl_", "").replace(".log", "")
                file_date = datetime.strptime(date_str, "%Y-%m-%d")
                
                if file_date < cutoff_date:
                    os.remove(log_file)
                    deleted += 1
                    self.logger.info(f"Удалён старый лог: {filename}")
            except (ValueError, OSError):
                pass
        
        if deleted > 0:
            self.logger.info(f"Удалено старых логов: {deleted}")
        else:
            self.logger.info(f"Старых логов не найдено (храним {self.KEEP_DAYS} дн.)")
    
    def get_logger(self):
        """Возвращает объект логгера для использования в других классах"""
        return self.logger
    
    def get_log_file(self):
        """Возвращает путь к файлу лога"""
        return self.log_file


# ============================================================
# КЛАСС 2: API КЛИЕНТ
# ============================================================

class APIClient:
    """
    Работа с внешним API.
    
    Возможности:
    - Формирование параметров запроса (даты, авторизация)
    - Получение данных из API
    - Парсинг вложенного поля passback_params
    """
    
    def __init__(self, logger):
        """
        Инициализация API клиента
        
        Args:
            logger: объект логгера из класса Logger
        """
        self.logger = logger
        self.raw_data = []          # Сырые данные из API
        self.cleaned_data = []      # Очищенные данные
        self.report_date = None     # Дата отчёта (вчера)
        self.params = None          # Параметры запроса
    
    def _prepare_params(self):
        """
        Внутренний метод: подготовка параметров для API запроса.
        Вычисляет start и end с учётом часового пояса.
        """
        hours = TIMEZONE_OFFSET
        yesterday = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
        self.report_date = yesterday
        
        # Формируем start и end с микросекундами
        start_dt = datetime.strptime(
            f'{yesterday} 00:00:00.000000', '%Y-%m-%d %H:%M:%S.%f'
        ) + timedelta(hours=hours)
        end_dt = datetime.strptime(
            f'{yesterday} 23:59:59.999999', '%Y-%m-%d %H:%M:%S.%f'
        ) + timedelta(hours=hours)
        
        self.params = {
            "client": API_CLIENT,
            "client_key": API_CLIENT_KEY,
            "start": start_dt.strftime("%Y-%m-%d %H:%M:%S.%f"),
            "end": end_dt.strftime("%Y-%m-%d %H:%M:%S.%f")
        }
        
        self.logger.info(f"Период: {yesterday}")
        self.logger.info(f"   Start: {self.params['start']}")
        self.logger.info(f"   End:   {self.params['end']}")
    
    def fetch_data(self):
        """
        Получение данных из API.
        
        Returns:
            bool: True если успешно, False при ошибке
        """
        self.logger.info("-" * 40)
        self.logger.info("ЭТАП 1: Запрос к API")
        self.logger.info("-" * 40)
        
        self._prepare_params()
        self.logger.info(f"URL: {API_URL}")
        
        try:
            response = requests.get(API_URL, params=self.params)
            self.logger.info(f"Статус: {response.status_code}")
            
            if response.status_code == 200:
                self.raw_data = response.json()
                self.logger.info(f"Получено записей: {len(self.raw_data)}")
                return True
            else:
                self.logger.error(f"Ошибка API! Статус: {response.status_code}")
                self.logger.error(f"   Ответ: {response.text[:200]}")
                return False
                
        except requests.exceptions.ConnectionError as e:
            self.logger.error(f"Ошибка подключения: {e}")
            return False
        except Exception as e:
            self.logger.error(f"Неизвестная ошибка: {e}")
            return False
    
    def parse_data(self):
        """
        Парсинг passback_params и формирование плоской структуры данных.
        
        Returns:
            bool: True если успешно, False при ошибке
        """
        self.logger.info("-" * 40)
        self.logger.info("Парсинг данных")
        self.logger.info("-" * 40)
        
        if not self.raw_data:
            self.logger.error("Нет данных для парсинга!")
            return False
        
        self.cleaned_data = []
        parse_errors = 0
        
        for i, record in enumerate(self.raw_data):
            try:
                # Базовые поля
                new_record = {
                    'user_id': record.get('lti_user_id'),
                    'is_correct': record.get('is_correct'),
                    'attempt_type': record.get('attempt_type'),
                    'created_at': record.get('created_at')
                }
                
                # Парсим строку passback_params в словарь
                passback_str = record.get('passback_params', '{}')
                try:
                    passback_dict = ast.literal_eval(passback_str)
                except:
                    passback_dict = {}
                    parse_errors += 1
                
                # Извлекаем нужные поля
                new_record['oauth_consumer_key'] = passback_dict.get('oauth_consumer_key', '')
                new_record['lis_result_sourcedid'] = passback_dict.get('lis_result_sourcedid', '')
                new_record['lis_outcome_service_url'] = passback_dict.get('lis_outcome_service_url', '')
                
                self.cleaned_data.append(new_record)
                
            except Exception as e:
                parse_errors += 1
                if parse_errors <= 3:
                    self.logger.warning(f"Ошибка в записи {i}: {e}")
        
        self.logger.info(f"Парсинг завершён!")
        self.logger.info(f"   Обработано: {len(self.cleaned_data)}")
        if parse_errors > 0:
            self.logger.warning(f"   Ошибок парсинга: {parse_errors}")
        
        return True
    
    def get_cleaned_data(self):
        """Возвращает очищенные данные"""
        return self.cleaned_data
    
    def get_report_date(self):
        """Возвращает дату отчёта"""
        return self.report_date


# ============================================================
# КЛАСС 3: БАЗА ДАННЫХ
# ============================================================

class Database:
    """
    Работа с PostgreSQL.
    
    Возможности:
    - Подключение к базе данных
    - Создание таблицы user_attempts
    - Загрузка данных
    - Агрегация статистики
    """
    
    TABLE_NAME = "user_attempts"
    
    def __init__(self, logger):
        """
        Инициализация объекта базы данных
        
        Args:
            logger: объект логгера
        """
        self.logger = logger
        self.conn = None            # Соединение с БД
        self.cur = None             # Курсор для запросов
        self.inserted_count = 0     # Количество вставленных записей
        self.stats = {}             # Статистика после агрегации
        self.hourly_data = []       # Почасовые данные
        self.top_users = []         # Топ пользователей
    
    def connect(self):
        """
        Установка соединения с PostgreSQL.
        
        Returns:
            bool: True если успешно, False при ошибке
        """
        try:
            self.conn = psycopg2.connect(
                dbname=DB_NAME,
                user=DB_USER,
                password=DB_PASSWORD,
                host=DB_HOST,
                port=DB_PORT
            )
            self.cur = self.conn.cursor()
            self.logger.info(f"Подключение к БД '{DB_NAME}' успешно!")
            return True
        except psycopg2.Error as e:
            self.logger.error(f"Ошибка подключения к БД: {e}")
            return False
    
    def create_table(self):
        """
        Создание таблицы user_attempts (если не существует).
        
        Returns:
            bool: True если успешно, False при ошибке
        """
        create_sql = f"""
        CREATE TABLE IF NOT EXISTS {self.TABLE_NAME} (
            id SERIAL PRIMARY KEY,
            user_id VARCHAR(255),
            oauth_consumer_key VARCHAR(255),
            lis_result_sourcedid TEXT,
            lis_outcome_service_url TEXT,
            is_correct INTEGER,
            attempt_type VARCHAR(50),
            created_at TIMESTAMP
        );
        """
        try:
            self.cur.execute(create_sql)
            self.conn.commit()
            self.logger.info(f"Таблица '{self.TABLE_NAME}' готова!")
            return True
        except psycopg2.Error as e:
            self.logger.error(f"Ошибка создания таблицы: {e}")
            return False
    
    def load_data(self, data, report_date):
        """
        Загрузка данных в таблицу.
        Сначала удаляет старые данные за этот день, потом вставляет новые.
        
        Args:
            data: список словарей с данными
            report_date: дата отчёта (строка YYYY-MM-DD)
        
        Returns:
            bool: True если успешно, False при ошибке
        """
        self.logger.info("-" * 40)
        self.logger.info("ЭТАП 2: Загрузка в PostgreSQL")
        self.logger.info("-" * 40)
        
        if not data:
            self.logger.error("Нет данных для загрузки!")
            return False
        
        # Удаляем старые данные за этот день
        try:
            self.cur.execute(
                f"DELETE FROM {self.TABLE_NAME} WHERE DATE(created_at) = %s",
                (report_date,)
            )
            deleted = self.cur.rowcount
            self.conn.commit()
            self.logger.info(f"   Удалено старых записей за {report_date}: {deleted}")
        except Exception as e:
            self.logger.warning(f"Ошибка удаления старых данных: {e}")
        
        # Загружаем новые данные
        insert_sql = f"""
            INSERT INTO {self.TABLE_NAME} 
            (user_id, oauth_consumer_key, lis_result_sourcedid, lis_outcome_service_url, 
             is_correct, attempt_type, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """
        
        inserted = 0
        errors = 0
        
        for record in data:
            try:
                values = (
                    record.get('user_id'),
                    record.get('oauth_consumer_key'),
                    record.get('lis_result_sourcedid'),
                    record.get('lis_outcome_service_url'),
                    record.get('is_correct'),
                    record.get('attempt_type'),
                    record.get('created_at')
                )
                self.cur.execute(insert_sql, values)
                inserted += 1
                
                # Коммитим каждые 100 записей для производительности
                if inserted % 100 == 0:
                    self.conn.commit()
                    self.logger.info(f"   Загружено {inserted} записей...")
                    
            except Exception as e:
                errors += 1
                if errors <= 3:
                    self.logger.warning(f"Ошибка вставки: {e}")
        
        self.conn.commit()
        self.inserted_count = inserted
        
        self.logger.info(f"ЗАГРУЗКА ЗАВЕРШЕНА!")
        self.logger.info(f"   Успешно: {inserted}")
        self.logger.info(f"   Ошибок: {errors}")
        
        return True
    
    def aggregate(self, report_date):
        """
        Агрегация данных для отчёта.
        
        Args:
            report_date: дата отчёта (строка YYYY-MM-DD)
        
        Returns:
            bool: True если успешно, False при ошибке
        """
        self.logger.info("-" * 40)
        self.logger.info("ЭТАП 3: Агрегация данных")
        self.logger.info("-" * 40)
        
        try:
            # 1. Общая статистика
            self.cur.execute(f"""
                SELECT 
                    COUNT(*) as total,
                    SUM(CASE WHEN is_correct = 1 THEN 1 ELSE 0 END) as correct,
                    SUM(CASE WHEN is_correct = 0 THEN 1 ELSE 0 END) as incorrect,
                    SUM(CASE WHEN attempt_type = 'run' THEN 1 ELSE 0 END) as runs,
                    SUM(CASE WHEN attempt_type = 'submit' THEN 1 ELSE 0 END) as submits,
                    COUNT(DISTINCT user_id) as unique_users
                FROM {self.TABLE_NAME}
                WHERE DATE(created_at) = %s
            """, (report_date,))
            
            row = self.cur.fetchone()
            if row:
                total, correct, incorrect, runs, submits, unique_users = row
                self.stats = {
                    'total': total or 0,
                    'correct': correct or 0,
                    'incorrect': incorrect or 0,
                    'runs': runs or 0,
                    'submits': submits or 0,
                    'unique_users': unique_users or 0
                }
                
                self.logger.info(f"Общая статистика за {report_date}:")
                self.logger.info(f"   Всего: {self.stats['total']}")
                self.logger.info(f"   Уникальных: {self.stats['unique_users']}")
                self.logger.info(f"   Правильных: {self.stats['correct']}")
                self.logger.info(f"   Неправильных: {self.stats['incorrect']}")
                self.logger.info(f"   Запусков (run): {self.stats['runs']}")
                self.logger.info(f"   Отправок (submit): {self.stats['submits']}")
            
            # 2. Почасовая статистика
            self.cur.execute(f"""
                SELECT 
                    EXTRACT(HOUR FROM created_at) as hour,
                    COUNT(*) as attempts,
                    SUM(CASE WHEN is_correct = 1 THEN 1 ELSE 0 END) as correct,
                    COUNT(DISTINCT user_id) as users
                FROM {self.TABLE_NAME}
                WHERE DATE(created_at) = %s
                GROUP BY EXTRACT(HOUR FROM created_at)
                ORDER BY hour
            """, (report_date,))
            
            self.hourly_data = self.cur.fetchall()
            self.logger.info(f"Почасовая статистика: {len(self.hourly_data)} часов с активностью")
            
            # 3. Топ-10 пользователей
            self.cur.execute(f"""
                SELECT 
                    user_id,
                    COUNT(*) as attempts,
                    SUM(CASE WHEN is_correct = 1 THEN 1 ELSE 0 END) as correct,
                    SUM(CASE WHEN is_correct = 0 THEN 1 ELSE 0 END) as incorrect
                FROM {self.TABLE_NAME}
                WHERE DATE(created_at) = %s
                GROUP BY user_id
                ORDER BY attempts DESC
                LIMIT 10
            """, (report_date,))
            
            self.top_users = self.cur.fetchall()
            self.logger.info(f"Топ пользователей: {len(self.top_users)} записей")
            
            return True
            
        except Exception as e:
            self.logger.error(f"Ошибка агрегации: {e}")
            return False
    
    def close(self):
        """Закрытие соединения с БД"""
        if self.cur:
            self.cur.close()
        if self.conn:
            self.conn.close()
        self.logger.info("Соединение с БД закрыто")
    
    # Геттеры для получения данных после агрегации
    def get_stats(self):
        return self.stats
    
    def get_hourly_data(self):
        return self.hourly_data
    
    def get_top_users(self):
        return self.top_users
    
    def get_inserted_count(self):
        return self.inserted_count


# ============================================================
# КЛАСС 4: EXCEL ОТЧЁТ
# ============================================================

class ExcelReport:
    """
    Создание Excel-отчёта с форматированием.
    
    Возможности:
    - Формирование структуры отчёта
    - Применение стилей (жирный шрифт, цвета, автоширина)
    - Сохранение в файл report_YYYY-MM-DD.xlsx
    """
    
    def __init__(self, logger):
        """
        Инициализация объекта Excel отчёта
        
        Args:
            logger: объект логгера
        """
        self.logger = logger
        self.filename = None         # Имя файла
        self.sheets_data = []        # Данные для записи
        self.report_date = None      # Дата отчёта для названия листа
    
    def generate(self, report_date, stats, hourly_data, top_users):
        """
        Генерация данных для Excel отчёта.
        
        Args:
            report_date: дата отчёта
            stats: словарь с общей статистикой
            hourly_data: почасовые данные
            top_users: топ пользователей
        
        Returns:
            bool: True
        """
        self.logger.info("-" * 40)
        self.logger.info("Подготовка данных для отчёта")
        self.logger.info("-" * 40)
        
        # Сохраняем дату для использования в save()
        self.report_date = report_date
        
        # Убедимся, что папка reports существует
        if not os.path.exists('reports'):
            os.makedirs('reports')
        self.filename = f'reports/report_{report_date}.xlsx'
        
        total = stats.get('total', 0)
        unique = stats.get('unique_users', 0)
        correct = stats.get('correct', 0)
        incorrect = stats.get('incorrect', 0)
        runs = stats.get('runs', 0)
        submits = stats.get('submits', 0)
        
        self.sheets_data = []
        
        # Заголовок
        self.sheets_data.append([f"ОТЧЁТ ПО ПОПЫТКАМ ЗА {report_date}"])
        self.sheets_data.append([])
        
        # Блок 1: Общая статистика
        self.sheets_data.append(["ОБЩАЯ СТАТИСТИКА"])
        self.sheets_data.append(["Показатель", "Значение"])
        self.sheets_data.append(["Дата", str(report_date)])
        self.sheets_data.append(["Всего попыток", total])
        self.sheets_data.append(["Уникальных пользователей", unique])
        self.sheets_data.append(["Правильных ответов", correct])
        self.sheets_data.append(["Неправильных ответов", incorrect])
        self.sheets_data.append(["Запусков кода (run)", runs])
        self.sheets_data.append(["Отправок на проверку (submit)", submits])
        
        if submits > 0:
            success_rate = (correct / submits) * 100
            self.sheets_data.append(["Успешность", f"{success_rate:.1f}%"])
        
        self.sheets_data.append([])
        
        # Блок 2: Почасовая статистика
        self.sheets_data.append(["СТАТИСТИКА ПО ЧАСАМ"])
        self.sheets_data.append(["Час", "Всего попыток", "Правильных", "Уникальных пользователей"])
        
        for hour, attempts, correct_h, users in hourly_data:
            self.sheets_data.append([f"{int(hour):02d}:00", attempts, correct_h, users])
        
        self.sheets_data.append([])
        
        # Блок 3: Топ пользователей
        self.sheets_data.append(["ТОП-10 ПОЛЬЗОВАТЕЛЕЙ ПО ПОПЫТКАМ"])
        self.sheets_data.append(["User ID", "Попыток", "Правильных", "Неправильных"])
        
        for user_id, attempts, correct_u, incorrect_u in top_users:
            self.sheets_data.append([user_id, attempts, correct_u, incorrect_u])
        
        self.logger.info(f"Данные подготовлены! ({len(self.sheets_data)} строк)")
        return True
    
    def save(self):
        """
        Сохранение Excel файла с форматированием.
        
        Returns:
            bool: True если успешно, False при ошибке
        """
        self.logger.info("-" * 40)
        self.logger.info("Сохранение Excel файла")
        self.logger.info("-" * 40)
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Создаём безопасное название листа (без слешей)
            if self.report_date:
                safe_date = self.report_date.replace('-', '_')
                ws.title = f"Отчёт_{safe_date}"[:31]  # Excel ограничивает 31 символом
            else:
                ws.title = "Статистика попыток"
            
            # Записываем данные и запоминаем строки с заголовками
            title_rows = []
            current_row = 1
            
            for row_data in self.sheets_data:
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=value)
                    if value and isinstance(value, str) and value.isupper():
                        title_rows.append(current_row)
                current_row += 1
            
            # Форматирование главного заголовка
            main_title = ws.cell(row=1, column=1)
            main_title.font = Font(bold=True, size=14, color="1F4E79")
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
            
            # Форматирование заголовков разделов
            header_font = Font(bold=True, size=11, color="1F4E79")
            header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            for row_num in title_rows:
                for col_idx in range(1, 5):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
            
            # Автоширина колонок
            for col_idx in range(1, 6):
                max_width = 10
                for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value:
                            width = len(str(cell.value)) * 1.2
                            max_width = min(max(max_width, width), 50)
                ws.column_dimensions[get_column_letter(col_idx)].width = max_width
            
            wb.save(self.filename)
            
            file_size = os.path.getsize(self.filename)
            self.logger.info(f"Excel файл сохранён!")
            self.logger.info(f"   Имя: {self.filename}")
            self.logger.info(f"   Размер: {file_size:,} байт ({file_size/1024:.1f} KB)")
            
            return True
            
        except Exception as e:
            self.logger.error(f"Ошибка создания Excel: {e}")
            return False
    
    def get_filename(self):
        """Возвращает имя созданного файла"""
        return self.filename


# ============================================================
# КЛАСС 5: EMAIL ОТПРАВИТЕЛЬ
# ============================================================

class EmailSender:
    """
    Отправка письма с отчётом.
    
    Возможности:
    - Формирование письма с HTML-телом
    - Прикрепление Excel файла
    - Отправка через SMTP (поддерживает TLS и SSL)
    """
    
    def __init__(self, logger):
        """
        Инициализация отправителя
        
        Args:
            logger: объект логгера
        """
        self.logger = logger
        self.sent = False
    
    def send(self, report_date, stats, filename):
        """
        Отправка письма с вложением.
        
        Args:
            report_date: дата отчёта
            stats: словарь со статистикой
            filename: имя файла для прикрепления
        
        Returns:
            bool: True если отправлено, False при ошибке
        """
        self.logger.info("-" * 40)
        self.logger.info("ЭТАП 4: Отправка Email")
        self.logger.info("-" * 40)
        
        # Проверяем существование файла
        if not os.path.exists(filename):
            self.logger.error(f"Файл {filename} не найден!")
            return False
        
        file_size = os.path.getsize(filename)
        self.logger.info(f"Файл: {filename} ({file_size/1024:.1f} KB)")
        self.logger.info(f"От: {EMAIL_SENDER}")
        self.logger.info(f"Кому: {EMAIL_RECEIVER}")
        
        total = stats.get('total', 0)
        unique = stats.get('unique_users', 0)
        correct = stats.get('correct', 0)
        incorrect = stats.get('incorrect', 0)
        runs = stats.get('runs', 0)
        
        # Создаём письмо
        msg = MIMEMultipart()
        msg['From'] = EMAIL_SENDER
        msg['To'] = EMAIL_RECEIVER
        msg['Subject'] = f"Отчёт по попыткам за {report_date}"
        
        # HTML тело письма
        body = f"""
        <html>
        <body>
            <h2>Ежедневный отчёт по попыткам</h2>
            <p><b>Дата:</b> {report_date}</p>
            <h3>Основные показатели:</h3>
            <ul>
                <li><b>Всего попыток:</b> {total}</li>
                <li><b>Уникальных пользователей:</b> {unique}</li>
                <li><b>Правильных:</b> {correct}</li>
                <li><b>Неправильных:</b> {incorrect}</li>
                <li><b>Запусков (run):</b> {runs}</li>
            </ul>
            <p>Подробная статистика — в прикреплённом файле <b>{os.path.basename(filename)}</b>.</p>
            <hr>
            <p><i>Письмо отправлено автоматически из ETL-системы.<br>
            По вопросам обращаться к стажёру-разработчику.</i></p>
        </body>
        </html>
        """
        msg.attach(MIMEText(body, 'html', 'utf-8'))
        
        # Прикрепляем Excel файл
        with open(filename, 'rb') as f:
            part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(filename)}"')
            msg.attach(part)
        
        self.logger.info(f"Подключение к {SMTP_SERVER}:{SMTP_PORT}...")
        
        try:
            # Определяем тип подключения по порту
            if SMTP_PORT == 465:
                # SSL подключение
                context = ssl.create_default_context()
                with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT, context=context, timeout=30) as server:
                    self.logger.info("Авторизация...")
                    server.login(EMAIL_SENDER, EMAIL_PASSWORD)
                    self.logger.info("Отправка...")
                    server.send_message(msg)
            else:
                # TLS подключение (порт 587)
                with smtplib.SMTP(SMTP_SERVER, SMTP_PORT, timeout=30) as server:
                    server.ehlo()
                    server.starttls()
                    server.ehlo()
                    self.logger.info("Авторизация...")
                    server.login(EMAIL_SENDER, EMAIL_PASSWORD)
                    self.logger.info("Отправка...")
                    server.send_message(msg)
            
            self.logger.info("ПИСЬМО УСПЕШНО ОТПРАВЛЕНО!")
            self.logger.info(f"   Получатель: {EMAIL_RECEIVER}")
            self.sent = True
            return True
            
        except smtplib.SMTPAuthenticationError as e:
            self.logger.error(f"Ошибка авторизации: {e}")
            self.logger.error("   Проверь логин и пароль в config.py")
            return False
        except Exception as e:
            self.logger.error(f"Ошибка отправки: {e}")
            return False
    
    def is_sent(self):
        """Возвращает статус отправки"""
        return self.sent


# ============================================================
# КЛАСС 6: ETL ПРОЦЕСС (ГЛАВНЫЙ КООРДИНАТОР)
# ============================================================

class ETLProcess:
    """
    Главный класс, управляющий всем ETL процессом.
    
    Координирует работу всех компонентов:
    - Logger
    - APIClient
    - Database
    - ExcelReport
    - EmailSender
    """
    
    def __init__(self):
        """Инициализация всех компонентов системы"""
        self.logger_manager = Logger()
        self.logger = self.logger_manager.get_logger()
        
        # Создаём компоненты, передавая им логгер
        self.api = APIClient(self.logger)
        self.db = Database(self.logger)
        self.excel = ExcelReport(self.logger)
        self.email = EmailSender(self.logger)
        
        # Отслеживание успешности каждого этапа
        self.success = {
            'api_fetch': False,
            'api_parse': False,
            'db_connect': False,
            'db_load': False,
            'db_aggregate': False,
            'excel_generate': False,
            'excel_save': False,
            'email_send': False
        }
    
    def run(self):
        """
        Запуск ETL процесса.
        Выполняет все этапы последовательно.
        При ошибке на любом этапе — останавливается.
        """
        try:
            # ========== ЭТАП 1: API ==========
            if not self.api.fetch_data():
                self._print_summary()
                return
            self.success['api_fetch'] = True
            
            if not self.api.parse_data():
                self._print_summary()
                return
            self.success['api_parse'] = True
            
            data = self.api.get_cleaned_data()
            report_date = self.api.get_report_date()
            
            # ========== ЭТАП 2: База данных ==========
            if not self.db.connect():
                self._print_summary()
                return
            self.success['db_connect'] = True
            
            if not self.db.create_table():
                self._print_summary()
                return
            
            if not self.db.load_data(data, report_date):
                self._print_summary()
                return
            self.success['db_load'] = True
            
            # ========== ЭТАП 3: Агрегация ==========
            if not self.db.aggregate(report_date):
                self._print_summary()
                return
            self.success['db_aggregate'] = True
            
            stats = self.db.get_stats()
            hourly = self.db.get_hourly_data()
            top = self.db.get_top_users()
            
            # ========== ЭТАП 4: Excel отчёт ==========
            if not self.excel.generate(report_date, stats, hourly, top):
                self._print_summary()
                return
            self.success['excel_generate'] = True
            
            if not self.excel.save():
                self._print_summary()
                return
            self.success['excel_save'] = True
            
            filename = self.excel.get_filename()
            
            # ========== ЭТАП 5: Email ==========
            if self.email.send(report_date, stats, filename):
                self.success['email_send'] = True
            
            # Закрываем соединение с БД
            self.db.close()
            
            # Выводим итоги
            self._print_summary()
            
        except Exception as e:
            self.logger.error(f"КРИТИЧЕСКАЯ ОШИБКА: {e}")
            self._print_summary()
    
    def _print_summary(self):
        """Вывод итогов выполнения в лог и консоль"""
        self.logger.info("=" * 50)
        self.logger.info("ETL-ПРОЦЕСС ЗАВЕРШЁН")
        self.logger.info("=" * 50)
        
        self.logger.info("ИТОГИ ВЫПОЛНЕНИЯ:")
        self.logger.info(f"   API получение:  {'[OK]' if self.success['api_fetch'] else '[FAIL]'}")
        self.logger.info(f"   API парсинг:    {'[OK]' if self.success['api_parse'] else '[FAIL]'}")
        self.logger.info(f"   БД подключение: {'[OK]' if self.success['db_connect'] else '[FAIL]'}")
        self.logger.info(f"   БД загрузка:    {'[OK]' if self.success['db_load'] else '[FAIL]'} ({self.db.get_inserted_count()} записей)")
        self.logger.info(f"   Агрегация:      {'[OK]' if self.success['db_aggregate'] else '[FAIL]'}")
        self.logger.info(f"   Excel генерация:{'[OK]' if self.success['excel_generate'] else '[FAIL]'}")
        self.logger.info(f"   Excel сохранение:{'[OK]' if self.success['excel_save'] else '[FAIL]'}")
        self.logger.info(f"   Email отправка: {'[OK]' if self.success['email_send'] else '[FAIL]'}")
        
        self.logger.info(f"\nЛог-файл: {self.logger_manager.get_log_file()}")
        self.logger.info(f"Excel-файл: {self.excel.get_filename() or 'не создан'}")
        
        # Вывод в консоль
        print("\n" + "=" * 50)
        print("ETL-ПРОЦЕСС ЗАВЕРШЁН!")
        print("=" * 50)
        print(f"Лог: {self.logger_manager.get_log_file()}")
        print(f"Отчёт: {self.excel.get_filename() or 'не создан'}")
        print(f"Письмо: {'[OK] Отправлено' if self.success['email_send'] else '[FAIL] Не отправлено'}")


# ============================================================
# ТОЧКА ВХОДА
# ============================================================

if __name__ == "__main__":
    """
    Запуск ETL-системы.
    Выполняется только при прямом запуске скрипта,
    не выполняется при импорте как модуля.
    """
    etl = ETLProcess()
    etl.run()